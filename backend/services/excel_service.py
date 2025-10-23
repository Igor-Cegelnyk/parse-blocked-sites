from io import BytesIO
from typing import Any, Type, TYPE_CHECKING, TypeVar, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse

from backend.models import Base

if TYPE_CHECKING:
    from pydantic import BaseModel

ModelType = TypeVar("ModelType", bound=Base)


class ExcelExportService:

    async def export_excel(
        self,
        data: List[ModelType],
        schema: Type["BaseModel"],
        filename: str | None = None,
    ) -> StreamingResponse:
        records = self._filter_data(data, schema)
        workbook = self._generate_excel(records, schema)
        return self._create_streaming_response(workbook, filename)

    @staticmethod
    def _filter_data(
        data: List[ModelType],
        schema: Type["BaseModel"],
    ) -> list[dict[str, Any]]:
        return (
            [schema.model_validate(item).model_dump(mode="json") for item in data]
            if data
            else []
        )

    @staticmethod
    def _generate_excel(
        data: list[dict[str, Any]],
        schema: Type["BaseModel"],
    ) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active

        # head
        fields = [
            (name, field.description or name)
            for name, field in schema.model_fields.items()
        ]
        for col_idx, (_, description) in enumerate(fields, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=description)
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(
                len(description) + 2, 12
            )

        # data
        for row_idx, item in enumerate(data, start=2):
            for col_idx, (field_name, _) in enumerate(fields, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=item[field_name])

        return workbook

    @staticmethod
    def _create_streaming_response(
        workbook: Workbook,
        filename: str | None,
    ) -> StreamingResponse:
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
